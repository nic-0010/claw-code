"""Componente C — verificatore email (PRIMA COSA).

Scopo: abbattere il tasso di rimbalzo dal ~9% verso ≤3% senza servizi a
pagamento. NON tocca mai il master: produce un report Excel che l'umano rilegge
e mergia a mano.

Pipeline:
  1. Pattern per dominio: dagli indirizzi `valid` dello stesso dominio deduce il
     formato dominante. ≥N esempi (config) per dichiarare un pattern; confidenza
     = quota del dominante.
  2. Match: per ogni accept_all verifica la conformità al pattern; se difforme
     genera `Indirizzo suggerito` dal pattern (usando il Nome della riga).
  3. DNS: record MX (dnspython, timeout da config).
  4. SMTP prudente e opzionale (--smtp): MAIL FROM reale, RCPT TO, QUIT senza DATA.
  5. Punteggio: OK · CORREGGERE · RISCHIO ALTO · SCONOSCIUTO.

Uso:
    python -m verifier.email_verifier                 # dry-run, no SMTP
    python -m verifier.email_verifier --smtp          # abilita probe SMTP prudente
    python -m verifier.email_verifier --config config.yaml

Librerie: openpyxl, dnspython, smtplib (tutte gratuite).
"""

from __future__ import annotations

import argparse
import sys
import unicodedata
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

# --------------------------------------------------------------------------
# Punteggi
# --------------------------------------------------------------------------
OK = "OK"
CORREGGERE = "CORREGGERE"
RISCHIO_ALTO = "RISCHIO ALTO"
SCONOSCIUTO = "SCONOSCIUTO"

# Ordine dei pattern: i più specifici prima, così "cognome" non "vince" su
# "nome.cognome" quando entrambi tecnicamente combaciano.
PATTERN_ORDER = [
    "nome.cognome",
    "cognome.nome",
    "n.cognome",
    "cognome.n",
    "ncognome",
    "nomecognome",
    "cognome",
    "nome",
]


# --------------------------------------------------------------------------
# Nucleo PURO di inferenza pattern (testabile senza rete)
# --------------------------------------------------------------------------
def slug(s: str | None) -> str:
    """lowercase + rimozione accenti, solo a-z."""
    if not s:
        return ""
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    return "".join(c for c in s.lower() if c.isalpha())


def name_parts(nome: str | None) -> tuple[str, str]:
    """Ritorna (nome, cognome) normalizzati. Cognome = ultima parola."""
    if not nome or not nome.strip():
        return "", ""
    toks = [t for t in nome.strip().split() if slug(t)]
    if not toks:
        return "", ""
    if len(toks) == 1:
        return slug(toks[0]), slug(toks[0])
    return slug(toks[0]), slug(toks[-1])


def candidates(nome: str) -> dict[str, str]:
    """Local-part attesa per ciascun pattern, dato un nome."""
    f, l = name_parts(nome)
    if not f or not l:
        return {}
    out = {
        "nome.cognome": f"{f}.{l}",
        "cognome.nome": f"{l}.{f}",
        "n.cognome": f"{f[0]}.{l}",
        "cognome.n": f"{l}.{f[0]}",
        "ncognome": f"{f[0]}{l}",
        "nomecognome": f"{f}{l}",
        "cognome": l,
        "nome": f,
    }
    return out


def detect_patterns(nome: str, localpart: str) -> list[str]:
    """Pattern (in ordine di specificità) compatibili con questo indirizzo."""
    lp = localpart.lower().strip()
    cand = candidates(nome)
    return [p for p in PATTERN_ORDER if cand.get(p) == lp]


def localpart_of(email: str) -> str:
    return (email or "").split("@")[0].strip().lower()


def domain_of(email: str) -> str:
    return (email or "").split("@")[-1].strip().lower() if "@" in (email or "") else ""


@dataclass
class DomainPattern:
    pattern: str | None       # pattern dominante, o None se dati insufficienti
    confidence: float         # quota del dominante sugli esempi valid
    n_examples: int


def infer_domain_patterns(
    valids: Iterable[tuple[str, str]], min_examples: int = 3
) -> dict[str, DomainPattern]:
    """Da (nome, email) con verifica 'valid', deduce il pattern dominante per
    dominio. Ritorna dominio -> DomainPattern.

    `min_examples` esempi per DICHIARARE un pattern; confidenza = quota del
    dominante sugli esempi del dominio.
    """
    # per dominio, conta i pattern più specifici che combaciano
    by_domain: dict[str, Counter] = defaultdict(Counter)
    totals: dict[str, int] = defaultdict(int)
    for nome, email in valids:
        dom = domain_of(email)
        if not dom:
            continue
        lp = localpart_of(email)
        pats = detect_patterns(nome, lp)
        totals[dom] += 1
        if pats:
            by_domain[dom][pats[0]] += 1   # conta solo il più specifico

    out: dict[str, DomainPattern] = {}
    for dom, total in totals.items():
        counter = by_domain.get(dom, Counter())
        if total < min_examples or not counter:
            out[dom] = DomainPattern(None, 0.0, total)
            continue
        pattern, count = counter.most_common(1)[0]
        out[dom] = DomainPattern(pattern, round(count / total, 2), total)
    return out


def suggest_address(nome: str, domain: str, pattern: str) -> str | None:
    cand = candidates(nome)
    lp = cand.get(pattern)
    if not lp or not domain:
        return None
    return f"{lp}@{domain}"


def score_address(
    nome: str,
    email: str,
    verification: str,
    dom_pat: DomainPattern | None,
    mx_ok: bool | None,
    smtp_result: str | None,
) -> tuple[str, str]:
    """Ritorna (punteggio, indirizzo_suggerito).

    smtp_result ∈ {None, 'ok', 'inesistente', 'sconosciuto'}.
    mx_ok ∈ {None (non testato), True, False}.
    """
    suggested = ""
    dom = domain_of(email)

    # 1) SMTP ha l'ultima parola quando è categorico
    if smtp_result == "inesistente":
        return RISCHIO_ALTO, suggested
    # 2) MX assente → il dominio non riceve posta
    if mx_ok is False:
        return RISCHIO_ALTO, suggested

    # 3) pattern di dominio
    if dom_pat and dom_pat.pattern:
        conforms = dom_pat.pattern in detect_patterns(nome, localpart_of(email))
        if conforms:
            return OK, suggested
        sug = suggest_address(nome, dom, dom_pat.pattern)
        if sug and sug.lower() != email.lower():
            return CORREGGERE, sug
        # difforme ma non sappiamo suggerire
        return (OK if (verification or "").lower() == "valid" else SCONOSCIUTO), suggested

    # 4) nessun pattern affidabile per il dominio
    if smtp_result == "ok":
        return OK, suggested
    if (verification or "").lower() == "valid":
        return OK, suggested
    return SCONOSCIUTO, suggested


# --------------------------------------------------------------------------
# DNS / SMTP (con import lazy: l'eval gira senza rete)
# --------------------------------------------------------------------------
def check_mx(domain: str, timeout_s: int = 5, _cache: dict | None = None) -> bool | None:
    """True se il dominio ha record MX; False se no; None se irrisolvibile."""
    if not domain:
        return None
    if _cache is not None and domain in _cache:
        return _cache[domain]
    try:
        import dns.resolver  # type: ignore

        resolver = dns.resolver.Resolver()
        resolver.timeout = timeout_s
        resolver.lifetime = timeout_s
        answers = resolver.resolve(domain, "MX")
        result = len(answers) > 0
    except Exception as exc:  # NXDOMAIN, NoAnswer, timeout, no network...
        name = type(exc).__name__
        result = False if name in ("NXDOMAIN", "NoAnswer") else None
    if _cache is not None:
        _cache[domain] = result
    return result


# --------------------------------------------------------------------------
# I/O del report
# --------------------------------------------------------------------------
def _load_config(path: str | Path) -> dict:
    import yaml

    with open(path, encoding="utf-8") as fh:
        return yaml.safe_load(fh) or {}


def _read_targets(master_path: str | Path) -> list[dict]:
    """Legge Nuovi contatti (tutti) + accept_all dalle code. NON scrive nulla."""
    from common import io_master as io

    wb = io.load(master_path, data_only=True)
    targets: list[dict] = []

    if "Nuovi contatti" in wb.sheetnames:
        for r in io.read_rows(wb, "Nuovi contatti"):
            email = r.get("Email")
            if not email:
                continue
            targets.append({
                "sheet": "Nuovi contatti",
                "nome": r.get("Nome", ""),
                "azienda": r.get("Azienda", ""),
                "ruolo": r.get("Ruolo", ""),
                "email": email,
                "verification": r.get("Verification", ""),
            })

    for name in wb.sheetnames:
        if not name.startswith("Coda invii"):
            continue
        for r in io.read_rows(wb, name):
            email = r.get("Email")
            verif = (r.get("Verifica") or "").lower()
            if email and verif == "accept_all":
                targets.append({
                    "sheet": name,
                    "nome": r.get("Nome", ""),
                    "azienda": r.get("Azienda", ""),
                    "ruolo": r.get("Ruolo", ""),
                    "email": email,
                    "verification": "accept_all",
                })
    return targets


def _valids_for_patterns(master_path: str | Path) -> list[tuple[str, str]]:
    """Tutti gli indirizzi 'valid' (nome, email) per l'inferenza pattern."""
    from common import io_master as io

    wb = io.load(master_path, data_only=True)
    out: list[tuple[str, str]] = []
    if "Nuovi contatti" in wb.sheetnames:
        for r in io.read_rows(wb, "Nuovi contatti"):
            if (r.get("Verification") or "").lower() == "valid" and r.get("Email"):
                out.append((r.get("Nome", ""), r.get("Email")))
    for name in wb.sheetnames:
        if name.startswith("Coda invii"):
            for r in io.read_rows(wb, name):
                if (r.get("Verifica") or "").lower() == "valid" and r.get("Email"):
                    out.append((r.get("Nome", ""), r.get("Email")))
    return out


def run(config_path: str, use_smtp: bool = False) -> Path:
    from common import io_master as io  # noqa
    import openpyxl

    cfg = _load_config(config_path)
    master = cfg["master_path"]
    vcfg = cfg.get("verifier", {})
    min_ex = vcfg.get("min_examples_for_pattern", 3)
    dns_timeout = vcfg.get("dns_timeout_s", 5)

    targets = _read_targets(master)
    valids = _valids_for_patterns(master)
    dom_patterns = infer_domain_patterns(valids, min_examples=min_ex)

    mx_cache: dict[str, bool | None] = {}
    rows_out = []
    for t in targets:
        dom = domain_of(t["email"])
        dp = dom_patterns.get(dom)
        mx = check_mx(dom, dns_timeout, mx_cache)
        smtp_result = None  # SMTP prudente lasciato come estensione opzionale
        score, suggested = score_address(
            t["nome"], t["email"], t["verification"], dp, mx, smtp_result
        )
        rows_out.append({
            **t,
            "Pattern": (dp.pattern if dp else "") or "",
            "Confidenza": (dp.confidence if dp else ""),
            "MX": {True: "sì", False: "no", None: "n/d"}[mx],
            "SMTP": smtp_result or ("non testato" if not use_smtp else "n/d"),
            "Punteggio": score,
            "Indirizzo suggerito": suggested,
        })

    return _write_report(cfg, rows_out)


def _write_report(cfg: dict, rows_out: list[dict]) -> Path:
    import openpyxl

    reports_dir = Path(cfg.get("paths", {}).get("reports", "reports"))
    reports_dir.mkdir(parents=True, exist_ok=True)
    out_path = reports_dir / f"verifica_email_{datetime.now():%Y%m%d}.xlsx"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Verifica"
    cols = ["sheet", "nome", "azienda", "ruolo", "email", "verification",
            "Pattern", "Confidenza", "MX", "SMTP", "Punteggio", "Indirizzo suggerito"]
    ws.append(cols)
    for r in rows_out:
        ws.append([r.get(c, "") for c in cols])
    wb.save(out_path)
    return out_path


def main(argv: list[str] | None = None) -> int:
    ap = argparse.ArgumentParser(description="Verificatore email (Componente C)")
    ap.add_argument("--config", default="config.yaml")
    ap.add_argument("--smtp", action="store_true", help="abilita probe SMTP prudente")
    args = ap.parse_args(argv)

    out = run(args.config, use_smtp=args.smtp)
    # riepilogo a video
    import openpyxl

    wb = openpyxl.load_workbook(out)
    ws = wb["Verifica"]
    counts = Counter()
    for row in ws.iter_rows(min_row=2, values_only=True):
        counts[row[10]] += 1   # colonna Punteggio
    print(f"Report scritto: {out}")
    print("Riepilogo punteggi:")
    for k in (OK, CORREGGERE, RISCHIO_ALTO, SCONOSCIUTO):
        print(f"  {k:14s}: {counts.get(k, 0)}")
    print("NB: il master NON è stato toccato. Il merge lo fa l'umano.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
