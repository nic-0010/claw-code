"""Test di io_master: backup, protezione colonne-formula, idempotenza, dry-run."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl
import pytest

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from common import io_master as io  # noqa: E402


def _make_master(tmp_path: Path) -> Path:
    """Crea un master minimale che imita il Registro invii reale."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro invii"
    headers = ["Data invio", "Email", "Ente/dominio", "Oggetto", "Stato",
               "Azione consigliata", "Risposta associata", "Preview risposta",
               "Gg lav da invio", "Variante"]
    for c, h in enumerate(headers, start=1):
        ws.cell(1, c, h)
    ws.cell(2, 2, "mario.rossi@esempio.it")
    ws.cell(2, 5, "Nessuna risposta")
    ws.cell(2, 9, "=IF(ISNUMBER(A2),1,\"\")")   # colonna I: formula read-only
    path = tmp_path / "master.xlsx"
    wb.save(path)
    return path


# --------------------------------------------------------------------------
# Backup
# --------------------------------------------------------------------------
def test_backup_crea_copia(tmp_path):
    master = _make_master(tmp_path)
    bdir = tmp_path / "backup"
    bkp = io.backup_master(master, bdir)
    assert bkp.exists()
    assert bkp.parent == bdir
    assert bkp.name.startswith("master_") and bkp.suffix == ".xlsx"
    # il backup è una copia byte-identica
    assert bkp.read_bytes() == master.read_bytes()


def test_backup_master_mancante(tmp_path):
    with pytest.raises(FileNotFoundError):
        io.backup_master(tmp_path / "nope.xlsx", tmp_path / "backup")


# --------------------------------------------------------------------------
# Protezione colonne formula
# --------------------------------------------------------------------------
def test_scrittura_colonna_formula_vietata(tmp_path):
    master = _make_master(tmp_path)
    wb = io.load(master)
    ws = wb["Registro invii"]
    with pytest.raises(io.FormulaCellError):
        io.set_cell(ws, 2, "I", 5)          # I = Gg lav da invio (formula)


def test_scrittura_su_cella_con_formula_vietata(tmp_path):
    master = _make_master(tmp_path)
    wb = io.load(master)
    ws = wb["Registro invii"]
    # anche se la colonna non fosse in blacklist, una formula reale è protetta
    with pytest.raises(io.FormulaCellError):
        io.set_cell(ws, 2, "I", "x")


def test_followup_colonne_formula():
    assert io.is_formula_column("Follow-up e riprese", "I")
    assert io.is_formula_column("Follow-up e riprese", "J")
    assert io.is_formula_column("Follow-up e riprese", "K")
    assert not io.is_formula_column("Follow-up e riprese", "H")   # Ultimo invio scrivibile


# --------------------------------------------------------------------------
# Idempotenza
# --------------------------------------------------------------------------
def test_set_cell_idempotente(tmp_path):
    master = _make_master(tmp_path)
    wb = io.load(master)
    ws = wb["Registro invii"]
    log = io.RunLog(component="test", apply=True)

    first = io.set_cell(ws, 2, "E", "Risposta positiva / referente", log=log)
    assert first is True
    assert len(log.writes) == 1

    # riscrivere lo stesso valore → no-op
    second = io.set_cell(ws, 2, "E", "Risposta positiva / referente", log=log)
    assert second is False
    assert len(log.writes) == 1
    assert len(log.skipped) == 1


def test_equal_none_vuoto():
    assert io._equal(None, "")
    assert io._equal(None, None)
    assert io._equal("  ciao ", "ciao")
    assert not io._equal("a", "b")


# --------------------------------------------------------------------------
# Dry-run vs apply
# --------------------------------------------------------------------------
def test_save_dry_run_non_scrive(tmp_path):
    master = _make_master(tmp_path)
    before = master.read_bytes()
    wb = io.load(master)
    ws = wb["Registro invii"]
    io.set_cell(ws, 2, "F", "nota di test")
    log = io.RunLog(component="test", apply=False)
    res = io.save(wb, master, apply=False, backup_dir=tmp_path / "backup", log=log)
    assert res is None
    assert master.read_bytes() == before          # file intatto
    assert any("DRY-RUN" in n for n in log.notes)


def test_save_apply_scrive_e_fa_backup(tmp_path):
    master = _make_master(tmp_path)
    wb = io.load(master)
    ws = wb["Registro invii"]
    io.set_cell(ws, 2, "F", "nota applicata")
    bdir = tmp_path / "backup"
    log = io.RunLog(component="test", apply=True)
    bkp = io.save(wb, master, apply=True, backup_dir=bdir, log=log)
    assert bkp is not None and bkp.exists()

    # rilettura: valore persistito
    wb2 = io.load(master)
    assert wb2["Registro invii"].cell(2, 6).value == "nota applicata"


# --------------------------------------------------------------------------
# read_rows
# --------------------------------------------------------------------------
def test_read_rows(tmp_path):
    master = _make_master(tmp_path)
    wb = io.load(master, data_only=False)
    rows = io.read_rows(wb, "Registro invii")
    assert len(rows) == 1
    assert rows[0]["Email"] == "mario.rossi@esempio.it"
    assert rows[0]["_row"] == 2
    assert rows[0]["Stato"] == "Nessuna risposta"


def test_runlog_save(tmp_path):
    log = io.RunLog(component="scanner", apply=False)
    log.notes.append("prova")
    p = log.save(tmp_path / "logs")
    assert p.exists()
    import json
    data = json.loads(p.read_text())
    assert data["component"] == "scanner"
    assert data["apply"] is False
