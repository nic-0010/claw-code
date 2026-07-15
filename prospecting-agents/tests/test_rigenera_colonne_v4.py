"""Test dello script di riallineamento colonne R/S V4 (scripts/rigenera_colonne_v4)."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts import rigenera_colonne_v4 as rig  # noqa: E402
from common import email_matrix as em  # noqa: E402

HEADERS = ["Priorità", "Variante test", "Segmento", "N", "Data", "Ora",
           "Azienda", "Nome", "Ruolo", "Email", "Verifica", "Confidenza",
           "Oggetto da usare", "Versione (V4 matrice)", "Stato", "Note",
           "Corpo mail", "Oggetto V4 (ruolo×società)", "Corpo V4 (ruolo×società)"]


def _make_master(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    wb.active.title = "Registro invii"      # foglio non-coda: ignorato
    wb["Registro invii"].append(["Data invio", "Email"])
    coda = wb.create_sheet("Coda invii dal 07-07")
    coda.append(HEADERS)
    # riga con R/S STALE
    coda.append([1, "", "Partecipata", 1, "", "", "GSE", "Marina Sacco",
                 "Director General", "marina.sacco@gse.it", "valid", 95,
                 "Ogg", "TAG_VECCHIO", "Da inviare", "",
                 "corpo A", "OGGETTO STALE", "chi ha un tempo scarsissimo... Detto ciò"])
    # riga senza Ruolo → saltata
    coda.append([2, "", "PA", 2, "", "", "Regione Lazio", "Mario Rossi", "",
                 "mario.rossi@regione.lazio.it", "valid", 90, "", "",
                 "Da inviare", "", "", "X", "Y"])
    path = tmp_path / "master.xlsx"
    wb.save(path)
    return path


def _cfg(tmp_path: Path, master: Path) -> dict:
    return {"master_path": str(master),
            "paths": {"backup": str(tmp_path / "backup"),
                      "logs": str(tmp_path / "logs")}}


def test_apply_riallinea_rs_a_build_email(tmp_path):
    master = _make_master(tmp_path)
    rig.rigenera(master, _cfg(tmp_path, master), apply=True)

    wb = openpyxl.load_workbook(master)
    ws = wb["Coda invii dal 07-07"]
    subj, body, tag = em.build_email("Marina Sacco", "GSE", "Director General",
                                     email="marina.sacco@gse.it")
    assert ws.cell(2, 18).value == subj          # R aggiornata
    assert ws.cell(2, 19).value == body          # S aggiornata
    assert ws.cell(2, 14).value == tag           # N (Versione) aggiornata
    # testo stale sparito, saluto genere-corretto presente
    assert "tempo scarsissimo" not in ws.cell(2, 19).value
    assert ws.cell(2, 19).value.startswith("Buongiorno Dott.ssa Sacco,")


def test_riga_senza_ruolo_saltata(tmp_path):
    master = _make_master(tmp_path)
    rig.rigenera(master, _cfg(tmp_path, master), apply=True)
    ws = openpyxl.load_workbook(master)["Coda invii dal 07-07"]
    assert ws.cell(3, 18).value == "X"           # invariata
    assert ws.cell(3, 19).value == "Y"


def test_dry_run_non_scrive_ma_conta(tmp_path):
    master = _make_master(tmp_path)
    before = master.read_bytes()
    log = rig.rigenera(master, _cfg(tmp_path, master), apply=False)
    assert master.read_bytes() == before         # file intatto
    assert len(log.writes) >= 3                   # R+S+N della riga valida
    assert not list((tmp_path / "backup").glob("*.xlsx"))  # nessun backup in dry-run


def test_apply_crea_backup(tmp_path):
    master = _make_master(tmp_path)
    log = rig.rigenera(master, _cfg(tmp_path, master), apply=True)
    assert log.backup_path and Path(log.backup_path).exists()


def test_idempotente(tmp_path):
    master = _make_master(tmp_path)
    rig.rigenera(master, _cfg(tmp_path, master), apply=True)
    log2 = rig.rigenera(master, _cfg(tmp_path, master), apply=True)
    assert len(log2.writes) == 0                  # già allineato → nessuna scrittura
