"""Test del queue builder (Componente B): batch, split A/B/C, cap, esclusioni,
render, output .eml, auto-rifornimento con V4, lead refill."""

from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from queue import queue_builder as qb  # noqa: E402
from queue import lead_refill as lr  # noqa: E402
from common import io_master as io  # noqa: E402

TODAY = datetime(2026, 7, 10)


# --------------------------------------------------------------------------
# Fixture: master sintetico completo
# --------------------------------------------------------------------------
def make_master(tmp_path: Path, n_coda: int = 15, n_nuovi: int = 10) -> Path:
    wb = openpyxl.Workbook()

    reg = wb.active
    reg.title = "Registro invii"
    reg.append(["Data invio", "Email", "Ente/dominio", "Oggetto", "Stato",
                "Azione consigliata", "Risposta associata", "Preview risposta",
                "Gg lav da invio", "Variante"])
    # un contatto con stato ≠ Nessuna risposta → escluso dai batch
    reg.append(["", "escluso.registro@enteX.it", "enteX.it", "", "Rifiuto / non interessato",
                "", "", "", "", ""])

    nr = wb.create_sheet("Non riscrivere")
    nr.append(["Score", "Nome", "Azienda", "", "", "", "", "", "Email"])
    nr.append([0, "Bruno Vietato", "Ente Blocco", "", "", "", "", "", ""])

    coda = wb.create_sheet("Coda invii dal 07-07")
    coda.append(qb.CODA_HEADERS)
    for i in range(1, n_coda + 1):
        coda.append([i, "", "PA / ente pubblico", i, "", "",
                     f"Ente{i}", f"Nome{i} Cognome{i}", "Manager",
                     f"nome{i}.cognome{i}@ente{i}.it", "valid", 90,
                     f"Oggetto {i}", "", "Da inviare", "",
                     f"Corpo A del contatto {i}, gentile [Cognome] di [Ente].",
                     "", ""])

    nc = wb.create_sheet("Nuovi contatti")
    nc.append(["Score", "Azione", "Segmento", "Persona", "Nome", "Azienda", "Ruolo",
               "Dipartimento", "Città", "Email", "Verification", "Confidence", "Nota"])
    for i in range(1, n_nuovi + 1):
        nc.append([100 + i, "NUOVO", "PA / ente pubblico", "Quadro",
                   f"Nuovo{i} Contatto{i}", f"NuovoEnte{i}", "Responsabile",
                   "Dip", "Roma", f"nuovo{i}.contatto{i}@nuovoente{i}.it",
                   "valid", 90, ""])

    fu = wb.create_sheet("Follow-up e riprese")
    fu.append(["Score", "Segmento", "Persona", "Nome", "Azienda", "Ruolo",
               "Email", "Ultimo invio", "Gg lavorativi", "Tipo azione", "Template"])
    # "Ultimo invio" con date reali rispetto a TODAY (10/07/2026); "Tipo azione"
    # lasciata VUOTA di proposito (come dopo un salvataggio openpyxl che azzera
    # la cache): il tipo deve essere ricalcolato in Python da "Ultimo invio".
    d_followup = datetime(2026, 6, 25)   # ~11 gg lavorativi fa → follow-up standard
    d_ripresa = datetime(2026, 5, 20)    # >25 gg lavorativi fa → ripresa
    for i in range(1, 13):
        ultimo = d_followup if i <= 8 else d_ripresa
        fu.append([50 + i, "PA", "Q", f"Fu{i} Cog{i}", f"FuEnte{i}", "Manager",
                   f"fu{i}@fuente{i}.it", ultimo, "", "", ""])

    tpl = wb.create_sheet("Template")
    tpl.append(["Template", "Oggetto", "Corpo"])
    tpl.append(["A · CONTROLLO — fredda PA / ente pubblico",
                "Breve confronto su previdenza", "Gentile [Cognome], corpo A per [Ente]."])
    tpl.append(["B · CHALLENGER V2 — fredda PA / ente pubblico",
                "Previdenza: cosa cambia dal 2026", "Gentile [Cognome], corpo B per [Ente]."])
    tpl.append(["SEQUENZA · F1 — bump nel thread (giorno 3-4)",
                "RE: [oggetto originario]", "Gentile [Cognome], torno sul tema."])
    tpl.append(["RIPRESA — contatti >25 gg",
                "Previdenza integrativa: cosa cambia nel 2026",
                "Gentile [Cognome], la ricontatto dopo qualche settimana."])

    path = tmp_path / "master.xlsx"
    wb.save(path)
    return path


CFG = {
    "caps": {"totale": 30, "fredde": 12, "follow_up": 10, "riprese": 5,
             "accept_all": 3, "stesso_dominio": 8},
    "domini_congelati": ["sace.it"],
    "soglia_scorta": 60,
    "sender_email": "io@me.it", "sender_name": "Io",
    "paths": {"backup": "backup", "logs": "logs", "bozze": "bozze"},
}


def _load(master):
    return io.load(master, data_only=True)


# --------------------------------------------------------------------------
# Batch: dimensioni, split, cap, esclusioni
# --------------------------------------------------------------------------
def test_batch_12_fredde_10_fu_5_riprese(tmp_path):
    master = make_master(tmp_path)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    fredde = [b for b in batch if b["tipo"] == "fredda"]
    fu = [b for b in batch if b["tipo"] == "follow-up"]
    rip = [b for b in batch if b["tipo"] == "ripresa"]
    assert len(fredde) == 12
    assert len(fu) == 8          # solo 8 follow-up maturi nel fixture
    assert len(rip) == 4         # 4 riprese nel fixture
    assert len(batch) <= 30


def test_followup_robusto_a_cache_formule_azzerate(tmp_path):
    """REGRESSIONE (bug di campo): con la col 'Tipo azione' come FORMULA la cui
    cache è azzerata (stato dopo un salvataggio openpyxl), il batch deve
    contenere comunque follow-up e riprese — calcolati da 'Ultimo invio'."""
    master = make_master(tmp_path)
    wb = openpyxl.load_workbook(master)
    ws = wb["Follow-up e riprese"]
    for r in range(2, ws.max_row + 1):
        ws.cell(r, 10, f'=IF(I{r}<4,"In attesa",IF(I{r}>25,"Ripresa","Follow-up"))')
        ws.cell(r, 9, f"=MAX(NETWORKDAYS(H{r},TODAY())-1,0)")
    wb.save(master)                                  # openpyxl → cache formule azzerata

    # data_only=True ora legge None per Tipo azione: senza il fix il batch
    # avrebbe 0 follow-up.
    reloaded = io.load(master, data_only=True)
    assert all(io.read_rows(reloaded, "Follow-up e riprese")[i].get("Tipo azione") is None
               for i in range(3))                    # cache davvero vuota
    batch = qb.select_batch(reloaded, CFG, today=TODAY)
    fu = [b for b in batch if b["tipo"] == "follow-up"]
    rip = [b for b in batch if b["tipo"] == "ripresa"]
    assert len(fu) == 8, [b["tipo"] for b in batch]
    assert len(rip) == 4


def test_split_abc_4_4_4(tmp_path):
    master = make_master(tmp_path)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    fredde = [b for b in batch if b["tipo"] == "fredda"]
    from collections import Counter
    c = Counter(b["variante"] for b in fredde)
    assert c == {"A": 4, "B": 4, "C": 4}


def test_variante_c_usa_matrice_v4(tmp_path):
    master = make_master(tmp_path)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    cs = [b for b in batch if b.get("variante") == "C"]
    assert cs
    for b in cs:
        assert b["versione"].startswith("C (")
        assert "secondo parere" in b["corpo"].lower()   # blocco anti-obiezione V4
        assert "⟦" not in b["corpo"]                     # niente segnaposto irrisolti


# Nuova apertura attesa per il cluster ruolo C (deve comparire nelle mail REALI).
NUOVA_APERTURA_C = (
    "chi ricopre un ruolo come il suo le decisioni importanti le ha già prese"
)
VECCHIA_APERTURA_C = "tempo scarsissimo"   # testo stale, non deve MAI comparire


def _master_con_rs_stale(tmp_path: Path) -> Path:
    """Coda con 3 fredde di cluster C (Direttore Generale) e colonne R/S del
    master già riempite con testo STALE. Con 3 fredde lo split è A/B/C su
    indici 0/1/2: la terza riga è la variante C."""
    wb = openpyxl.Workbook()
    wb.active.title = "Registro invii"
    wb["Registro invii"].append(["Data invio", "Email", "Ente/dominio", "Oggetto",
                                 "Stato", "Azione consigliata", "Risposta associata",
                                 "Preview risposta", "Gg lav da invio", "Variante"])
    coda = wb.create_sheet("Coda invii dal 07-07")
    coda.append(qb.CODA_HEADERS)
    stale_body = ("Gentile [Cognome], chi ha una responsabilità di vertice come "
                  "la sua ha, di solito, un tempo scarsissimo. Detto ciò, ...")
    for i in range(1, 4):
        coda.append([i, "", "Partecipata / grande ente", i, "", "",
                     "GSE", f"Nome{i} Cognome{i}", "Direttore Generale",
                     f"nome{i}.cognome{i}@gse.it", "valid", 95,
                     "Oggetto coda", "PARTECIPATA·C", "Da inviare", "",
                     "corpo A", "OGGETTO V4 STALE", stale_body])
    path = tmp_path / "master_stale.xlsx"
    wb.save(path)
    return path


def test_variante_c_ignora_colonne_rs_stale_del_master(tmp_path):
    """REGRESSIONE: il builder NON deve usare le colonne R/S del master (che
    possono essere stale). La C viene sempre rigenerata da build_email()."""
    master = _master_con_rs_stale(tmp_path)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    cs = [b for b in batch if b.get("variante") == "C"]
    assert len(cs) == 1
    corpo = cs[0]["corpo"]
    assert NUOVA_APERTURA_C in corpo               # nuova apertura presente
    assert VECCHIA_APERTURA_C not in corpo         # testo stale assente
    assert "Detto ciò" not in corpo


FIRMA_ATTESA = (
    "Nicolò Porru\n"
    "Wealth & Insurance Advisor — Generali Italia\n"
    "Tel +39 331 454 8168 · LinkedIn: linkedin.com/in/nicolò-porru"
)


def test_ogni_eml_del_batch_contiene_la_firma(tmp_path):
    """OGNI bozza generata (A, B e C — da master/Template o dalla matrice) deve
    contenere ESATTAMENTE la firma canonica."""
    from email import message_from_bytes

    master = make_master(tmp_path)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    out_dir = qb.write_outputs(batch, CFG, out_root=tmp_path / "bozze", today=TODAY)

    varianti = {b.get("variante") for b in batch}
    assert {"A", "B", "C"} <= varianti          # il batch copre tutte le varianti

    emls = sorted(out_dir.glob("*.eml"))
    assert len(emls) == len(batch)
    for eml in emls:
        payload = message_from_bytes(eml.read_bytes()).get_payload(decode=True).decode("utf-8")
        assert FIRMA_ATTESA in payload, eml.name
        assert "Wealth & Insurance Advisor" in payload
        assert "+39 331 454 8168" in payload
        assert "linkedin.com/in/nicolò-porru" in payload


def test_firma_non_duplicata_sulla_variante_c(tmp_path):
    """La C include già la firma da build_email(): non deve comparire due volte."""
    master = make_master(tmp_path)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    for b in batch:
        if b.get("variante") == "C":
            assert b["corpo"].count("Wealth & Insurance Advisor") == 1


def test_eml_reale_contiene_nuova_apertura_c(tmp_path):
    """Confronta l'output REALE (.eml scritto su disco) col testo atteso —
    non lo snapshot. Guardia contro il disallineamento matrice↔mail reale."""
    from email import message_from_bytes

    master = _master_con_rs_stale(tmp_path)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    out_dir = qb.write_outputs(batch, CFG, out_root=tmp_path / "bozze", today=TODAY)

    c_items = [b for b in batch if b.get("variante") == "C"]
    assert c_items
    c_email = c_items[0]["email"].lower().replace("@", "_").replace(".", "_")
    eml = next(p for p in out_dir.glob("*.eml") if c_email in p.name)

    msg = message_from_bytes(eml.read_bytes())
    payload = msg.get_payload(decode=True).decode("utf-8")
    assert NUOVA_APERTURA_C in payload             # la mail reale ha il nuovo testo
    assert VECCHIA_APERTURA_C not in payload
    assert "protezione del patrimonio" in payload  # tema dell'apertura C


def test_esclusione_registro_e_non_riscrivere(tmp_path):
    master = make_master(tmp_path, n_coda=3)
    wb = openpyxl.load_workbook(master)
    coda = wb["Coda invii dal 07-07"]
    # contatto presente nel Registro con stato ≠ Nessuna risposta
    coda.append([99, "", "PA", 99, "", "", "EnteX", "Tizio Caio", "Manager",
                 "escluso.registro@enteX.it", "valid", 90, "", "", "Da inviare",
                 "", "corpo", "", ""])
    # contatto in Non riscrivere (match Nome+Azienda)
    coda.append([98, "", "PA", 98, "", "", "Ente Blocco", "Bruno Vietato", "Manager",
                 "bruno.vietato@enteblocco.it", "valid", 90, "", "", "Da inviare",
                 "", "corpo", "", ""])
    wb.save(master)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    emails = {b["email"].lower() for b in batch}
    assert "escluso.registro@entex.it" not in emails
    assert "bruno.vietato@enteblocco.it" not in emails


def test_zero_contatti_in_blacklist_nel_batch(tmp_path):
    """Test automatico richiesto dalla spec (§B Done)."""
    master = make_master(tmp_path)
    wb = _load(master)
    batch = qb.select_batch(wb, CFG, today=TODAY)
    non_riscrivere, stato_email = qb.build_exclusions(wb)
    for b in batch:
        assert b["email"].lower() not in stato_email
        assert qb._norm_key(b["nome"], b["azienda"]) not in non_riscrivere


def test_cap_accept_all(tmp_path):
    master = make_master(tmp_path, n_coda=0)
    wb = openpyxl.load_workbook(master)
    coda = wb["Coda invii dal 07-07"]
    for i in range(1, 9):
        coda.append([i, "", "PA", i, "", "", f"E{i}", f"N{i} C{i}", "Manager",
                     f"n{i}.c{i}@e{i}.it", "accept_all", 60, "", "", "Da inviare",
                     "", "corpo", "", ""])
    wb.save(master)
    cfg = {**CFG, "caps": {**CFG["caps"], "fredde": 12}}
    batch = qb.select_batch(_load(master), cfg, today=TODAY)
    accept = [b for b in batch if b["tipo"] == "fredda" and b["fonte"].startswith("Coda")]
    assert len(accept) <= 3 + 12  # sanity
    n_accept = sum(1 for b in batch if b.get("verifica") == "accept_all")
    assert n_accept <= 3


def test_cap_stesso_dominio(tmp_path):
    master = make_master(tmp_path, n_coda=0, n_nuovi=0)
    wb = openpyxl.load_workbook(master)
    coda = wb["Coda invii dal 07-07"]
    for i in range(1, 13):
        coda.append([i, "", "PA", i, "", "", "StessoEnte", f"P{i} Q{i}", "Manager",
                     f"p{i}.q{i}@stessodominio.it", "valid", 90, "", "",
                     "Da inviare", "", "corpo", "", ""])
    wb.save(master)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    fredde = [b for b in batch if b["tipo"] == "fredda"]
    assert len(fredde) == 8      # cap stesso_dominio


def test_fallback_a_nuovi_contatti_quando_code_esaurite(tmp_path):
    master = make_master(tmp_path, n_coda=2, n_nuovi=20)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    fredde = [b for b in batch if b["tipo"] == "fredda"]
    assert len(fredde) == 12
    assert sum(1 for b in fredde if b["fonte"] == "Nuovi contatti") == 10


# --------------------------------------------------------------------------
# Render
# --------------------------------------------------------------------------
def test_applescript_draft_salva_mai_invia():
    item = {"oggetto": 'Ogg "citato"', "corpo": "riga1\nriga2 con \"virg\"",
            "email": "mario@ente.it"}
    script = qb.applescript_draft(item)
    assert "Microsoft Outlook" in script
    assert "save newMsg" in script          # persiste come bozza
    assert "send" not in script             # MAI invio
    # escaping: virgolette escapate, newline reale via linefeed
    assert '\\"citato\\"' in script
    assert "linefeed" in script
    assert "mario@ente.it" in script
    # nessun newline grezzo dentro il literal della property content
    content_line = [l for l in script.splitlines() if "plain text content" in l][0]
    assert "riga1" in content_line and "riga2" in content_line


def test_create_outlook_drafts_fallback_non_macos(monkeypatch):
    # forza ambiente non-macOS → available False, il chiamante userà i .eml
    monkeypatch.setattr(qb.sys, "platform", "linux")
    res = qb.create_outlook_drafts([{"email": "x@y.it", "oggetto": "o", "corpo": "c"}], CFG)
    assert res["available"] is False
    assert res["created"] == 0
    assert res["reason"]


def test_render_placeholders():
    out = qb.render_placeholders(
        "Gentile [Cognome] di [Ente], ci vediamo a [mese] il [GG] alle [HH:MM].",
        "Mario Rossi", "ACME", today=TODAY)
    assert "Gentile Rossi di ACME" in out
    assert "a luglio" in out
    assert "⟦GG⟧" in out and "⟦HH:MM⟧" in out    # irrisolti evidenziati


# --------------------------------------------------------------------------
# Output .eml + riepilogo
# --------------------------------------------------------------------------
def test_write_outputs_eml_e_riepilogo(tmp_path):
    master = make_master(tmp_path)
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    out_dir = qb.write_outputs(batch, CFG, out_root=tmp_path / "bozze", today=TODAY)
    emls = sorted(out_dir.glob("*.eml"))
    assert len(emls) == len(batch)
    # ogni .eml è RFC822 apribile
    from email import message_from_bytes
    msg = message_from_bytes(emls[0].read_bytes())
    assert msg["To"] and msg["Subject"] and msg.get_payload()
    assert (out_dir / "riepilogo.html").exists()
    html_text = (out_dir / "riepilogo.html").read_text(encoding="utf-8")
    assert f"{len(batch)} bozze" in html_text


# --------------------------------------------------------------------------
# Auto-rifornimento
# --------------------------------------------------------------------------
def test_refill_sotto_soglia_crea_coda_con_v4(tmp_path):
    master = make_master(tmp_path, n_nuovi=10)     # 10 valid < soglia 60
    res = qb.refill_queue(master, CFG, apply=True, today=TODAY)
    assert res["created"] is True
    wb = openpyxl.load_workbook(master)
    sheet = f"Coda invii dal {TODAY:%d-%m}"
    assert sheet in wb.sheetnames
    ws = wb[sheet]
    assert [c.value for c in ws[1]] == qb.CODA_HEADERS
    # R/S V4 compilate
    assert ws.cell(2, 18).value            # Oggetto V4
    assert "secondo parere" in str(ws.cell(2, 19).value).lower()
    # Nota flaggata sui contatti presi
    nc = wb["Nuovi contatti"]
    assert str(nc.cell(2, 13).value).startswith("In coda dal")


def test_refill_sopra_soglia_non_fa_nulla(tmp_path):
    master = make_master(tmp_path, n_nuovi=10)
    cfg = {**CFG, "soglia_scorta": 5}              # 10 valid ≥ 5
    res = qb.refill_queue(master, cfg, apply=True, today=TODAY)
    assert res["created"] is False


def test_refill_idempotente(tmp_path):
    master = make_master(tmp_path, n_nuovi=10)
    qb.refill_queue(master, CFG, apply=True, today=TODAY)
    res2 = qb.refill_queue(master, CFG, apply=True, today=TODAY)
    assert res2["created"] is False                # il foglio del giorno esiste già
    wb = openpyxl.load_workbook(master)
    assert sum(1 for s in wb.sheetnames
               if s == f"Coda invii dal {TODAY:%d-%m}") == 1


def test_refill_dry_run_non_scrive(tmp_path):
    master = make_master(tmp_path, n_nuovi=10)
    before = master.read_bytes()
    res = qb.refill_queue(master, CFG, apply=False, today=TODAY)
    assert res["created"] is True                  # avrebbe creato...
    assert master.read_bytes() == before           # ...ma non ha scritto


def test_simulazione_svuotamento_genera_coda_v4(tmp_path):
    """Spec §B Done: svuotando le code, il sistema genera da solo la coda
    successiva con le V4 pronte — nessuna mattina senza batch."""
    master = make_master(tmp_path, n_coda=0, n_nuovi=15)
    res = qb.refill_queue(master, CFG, apply=True, today=TODAY)
    assert res["created"] and res["n"] == 15
    batch = qb.select_batch(_load(master), CFG, today=TODAY)
    fredde = [b for b in batch if b["tipo"] == "fredda"]
    assert len(fredde) == 12                       # batch pieno dalla nuova coda


# --------------------------------------------------------------------------
# Catena completa dell'auto-rifornimento: coda → lead_refill automatico
# --------------------------------------------------------------------------
def _make_archive(tmp_path: Path, n: int = 80) -> Path:
    arch = openpyxl.Workbook()
    ws = arch.active
    ws.append(["Nome", "Azienda", "Ruolo", "Email", "Verification", "Segmento"])
    for i in range(1, n + 1):
        ws.append([f"Arch{i} Lead{i}", f"ArchEnte{i}", "Manager",
                   f"arch{i}.lead{i}@archente{i}.it", "valid", "PA"])
    p = tmp_path / "archivio.xlsx"
    arch.save(p)
    return p


def test_ensure_lead_stock_sotto_soglia_importa(tmp_path):
    """Spec §B punto 2: Nuovi contatti sotto soglia → lead_refill automatico."""
    master = make_master(tmp_path, n_nuovi=5)      # 5 valid < soglia 60
    cfg = {**CFG, "lead_archive": str(_make_archive(tmp_path))}
    res = qb.ensure_lead_stock(master, cfg, apply=True)
    assert res["triggered"] is True
    assert res["appended"] > 0
    assert res["n_valid"] >= 5 + res["appended"] - 1   # scorte risalite
    wb = openpyxl.load_workbook(master)
    nc = wb["Nuovi contatti"]
    assert "arch" in str(nc.cell(nc.max_row, 10).value)


def test_ensure_lead_stock_sopra_soglia_non_importa(tmp_path):
    master = make_master(tmp_path, n_nuovi=10)
    cfg = {**CFG, "soglia_scorta": 5,
           "lead_archive": str(_make_archive(tmp_path))}
    res = qb.ensure_lead_stock(master, cfg, apply=True)
    assert res["triggered"] is False


def test_ensure_lead_stock_senza_archivio_warning(tmp_path):
    master = make_master(tmp_path, n_nuovi=5)
    cfg = {**CFG, "lead_archive": str(tmp_path / "inesistente.xlsx")}
    res = qb.ensure_lead_stock(master, cfg, apply=True)
    assert res["triggered"] is False
    assert "archivio" in res["warning"]


def test_ensure_lead_stock_dry_run_non_scrive(tmp_path):
    master = make_master(tmp_path, n_nuovi=5)
    cfg = {**CFG, "lead_archive": str(_make_archive(tmp_path))}
    before = master.read_bytes()
    res = qb.ensure_lead_stock(master, cfg, apply=False)
    assert res["triggered"] is True
    assert master.read_bytes() == before


def test_catena_completa_svuotamento_coda_e_archivio(tmp_path):
    """Mai a secco: coda vuota → refill coda con V4; scorte esaurite dopo
    l'accodamento → lead_refill dall'archivio. Due stadi in un solo run."""
    master = make_master(tmp_path, n_coda=0, n_nuovi=15)
    cfg = {**CFG, "lead_archive": str(_make_archive(tmp_path, n=100))}
    res1 = qb.refill_queue(master, cfg, apply=True, today=TODAY)
    assert res1["created"] and res1["n"] == 15     # tutti accodati → scorte a 0
    res2 = qb.ensure_lead_stock(master, cfg, apply=True)
    assert res2["triggered"] is True
    assert res2["n_valid"] >= cfg["soglia_scorta"]  # scorte ricostituite


# --------------------------------------------------------------------------
# Lead refill dall'archivio
# --------------------------------------------------------------------------
def test_lead_refill_dedupe_e_append(tmp_path):
    master = make_master(tmp_path, n_nuovi=3)
    arch = openpyxl.Workbook()
    ws = arch.active
    ws.append(["Nome", "Azienda", "Ruolo", "Email", "Verification", "Segmento"])
    ws.append(["Nuovo1 Contatto1", "NuovoEnte1", "Manager",
               "nuovo1.contatto1@nuovoente1.it", "valid", "PA"])   # dup email
    ws.append(["Franca Neri", "Ente Nuovissimo", "Director General",
               "franca.neri@nuovissimo.it", "valid", "PA"])        # nuovo
    ws.append(["Bruno Vietato", "Ente Blocco", "Manager",
               "bruno.v@enteblocco.it", "valid", "PA"])            # dup Nome+Azienda
    arch_path = tmp_path / "archivio.xlsx"
    arch.save(arch_path)

    res = lr.refill(master, arch_path, CFG, apply=True, top_n=10)
    assert res["appended"] == 1
    wb = openpyxl.load_workbook(master)
    nc = wb["Nuovi contatti"]
    last = nc.max_row
    assert nc.cell(last, 10).value == "franca.neri@nuovissimo.it"
    assert nc.cell(last, 2).value == "NUOVO CONTATTO"
    assert (nc.cell(last, 1).value or 0) >= 150    # score da Director General
