"""Test di scripts/log_invii: registrazione batch nel Registro + Ultimo invio."""

from __future__ import annotations

import sys
from datetime import date, datetime
from email.message import EmailMessage
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scripts import log_invii as li  # noqa: E402
from common import io_master as io  # noqa: E402

DATA = date(2026, 7, 16)


def _write_eml(d: Path, idx: int, email: str, variante: str, versione: str,
               oggetto: str = "Oggetto test"):
    msg = EmailMessage()
    msg["From"] = "Io <io@me.it>"
    msg["To"] = email
    msg["Subject"] = oggetto
    msg["X-Prospecting-Variante"] = variante
    msg["X-Prospecting-Versione"] = versione
    msg.set_content("corpo")
    (d / f"{idx:02d}_{variante}_{email.replace('@','_')}.eml").write_bytes(bytes(msg))


def _make_bozze(tmp_path: Path) -> Path:
    d = tmp_path / "bozze" / "20260716"
    d.mkdir(parents=True)
    _write_eml(d, 1, "mario.rossi@ente1.it", "A", "A")
    _write_eml(d, 2, "lucia.bianchi@ente2.it", "C", "C (ORGINT·C)")
    _write_eml(d, 3, "fu1@fuente.it", "—", "F1")            # follow-up
    _write_eml(d, 4, "rip1@ripente.it", "—", "RIPRESA")     # ripresa
    return d


def _make_master(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    reg = wb.active
    reg.title = "Registro invii"
    reg.append(["Data invio", "Email", "Ente/dominio", "Oggetto", "Stato",
                "Azione consigliata", "Risposta associata", "Preview risposta",
                "Gg lav da invio", "Variante"])
    fu = wb.create_sheet("Follow-up e riprese")
    fu.append(["Score", "Segmento", "Persona", "Nome", "Azienda", "Ruolo",
               "Email", "Ultimo invio", "Gg lavorativi", "Tipo azione", "Template"])
    fu.append([100, "PA", "Q", "Fu Uno", "FuEnte", "Manager", "fu1@fuente.it",
               "", "", "", ""])
    fu.append([90, "PA", "Q", "Rip Uno", "RipEnte", "Manager", "rip1@ripente.it",
               "", "", "", ""])
    p = tmp_path / "master.xlsx"
    wb.save(p)
    return p


def _cfg(tmp_path):
    return {"paths": {"backup": str(tmp_path / "b"), "logs": str(tmp_path / "l")}}


def test_apply_registra_batch(tmp_path):
    master = _make_master(tmp_path)
    bozze = _make_bozze(tmp_path)
    res = li.log_sends(master, bozze, _cfg(tmp_path), apply=True)
    assert res["logged"] == 4
    reg = openpyxl.load_workbook(master)["Registro invii"]
    # 1 header + 4 righe
    assert reg.max_row == 5
    row = {reg.cell(1, c).value: reg.cell(2, c).value for c in range(1, 11)}
    assert row["Email"] == "mario.rossi@ente1.it"
    assert row["Ente/dominio"] == "ente1.it"
    assert row["Stato"] == "Nessuna risposta"
    assert row["Data invio"] == datetime(2026, 7, 16)
    assert row["Variante"] == "A"


def test_variante_c_registrata_e_dash_vuota(tmp_path):
    master = _make_master(tmp_path)
    bozze = _make_bozze(tmp_path)
    li.log_sends(master, bozze, _cfg(tmp_path), apply=True)
    reg = openpyxl.load_workbook(master)["Registro invii"]
    vals = {reg.cell(r, 2).value: reg.cell(r, 10).value for r in range(2, reg.max_row + 1)}
    assert vals["lucia.bianchi@ente2.it"] == "C"
    assert vals["fu1@fuente.it"] in (None, "")            # variante "—" → vuota


def test_followup_ultimo_invio_aggiornato(tmp_path):
    master = _make_master(tmp_path)
    bozze = _make_bozze(tmp_path)
    res = li.log_sends(master, bozze, _cfg(tmp_path), apply=True)
    assert res["followup_updated"] == 2
    fu = openpyxl.load_workbook(master)["Follow-up e riprese"]
    ultimo = {fu.cell(r, 7).value: fu.cell(r, 8).value for r in range(2, fu.max_row + 1)}
    assert ultimo["fu1@fuente.it"] == datetime(2026, 7, 16)
    assert ultimo["rip1@ripente.it"] == datetime(2026, 7, 16)


def test_idempotente(tmp_path):
    master = _make_master(tmp_path)
    bozze = _make_bozze(tmp_path)
    li.log_sends(master, bozze, _cfg(tmp_path), apply=True)
    res2 = li.log_sends(master, bozze, _cfg(tmp_path), apply=True)
    assert res2["logged"] == 0 and res2["skipped"] == 4
    reg = openpyxl.load_workbook(master)["Registro invii"]
    assert reg.max_row == 5                               # nessuna riga doppia


def test_escludi_non_inviate(tmp_path):
    master = _make_master(tmp_path)
    bozze = _make_bozze(tmp_path)
    res = li.log_sends(master, bozze, _cfg(tmp_path), apply=True,
                       escludi={"mario.rossi@ente1.it"})
    assert res["logged"] == 3 and res["excluded"] == 1
    reg = openpyxl.load_workbook(master)["Registro invii"]
    emails = {reg.cell(r, 2).value for r in range(2, reg.max_row + 1)}
    assert "mario.rossi@ente1.it" not in emails


def test_dry_run_non_scrive(tmp_path):
    master = _make_master(tmp_path)
    bozze = _make_bozze(tmp_path)
    before = master.read_bytes()
    res = li.log_sends(master, bozze, _cfg(tmp_path), apply=False)
    assert res["logged"] == 4                             # avrebbe registrato...
    assert master.read_bytes() == before                 # ...ma non ha scritto


def test_data_dal_nome_cartella(tmp_path):
    master = _make_master(tmp_path)
    bozze = _make_bozze(tmp_path)
    res = li.log_sends(master, bozze, _cfg(tmp_path), apply=True,
                       today=date(2030, 1, 1))            # today diverso dalla cartella
    assert res["data"] == date(2026, 7, 16)               # vince il nome cartella
