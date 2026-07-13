"""Test del reply scanner (Componente A): deterministico, euristiche,
scritture idempotenti nel Registro, colonne proibite."""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from scanner import reply_scanner as rs  # noqa: E402
from common import io_master as io  # noqa: E402


# --------------------------------------------------------------------------
# Classificazione deterministica
# --------------------------------------------------------------------------
def test_bounce_da_mittente():
    assert rs.classify_deterministic("postmaster@x.it", "qualsiasi", "") == rs.BOUNCE
    assert rs.classify_deterministic("MAILER-DAEMON@y.net", "", "") == rs.BOUNCE


def test_bounce_da_oggetto():
    assert rs.classify_deterministic("a@b.it", "Undeliverable: test", "") == rs.BOUNCE
    assert rs.classify_deterministic("a@b.it", "Mancato recapito", "") == rs.BOUNCE


def test_oof():
    assert rs.classify_deterministic("a@b.it", "Automatic reply: X", "") == rs.AUTO_REPLY
    assert rs.classify_deterministic("a@b.it", "Risposta automatica: X", "") == rs.AUTO_REPLY
    assert rs.classify_deterministic("a@b.it", "RE: X", "Sono assente dal 10 agosto") == rs.AUTO_REPLY


def test_non_deterministico_passa_al_modello():
    assert rs.classify_deterministic("a@b.it", "RE: proposta", "mi interessa molto") is None


def test_bounce_ndr_exchange_italiano():
    """Pattern dai casi reali: NDR Office365/Exchange in italiano."""
    assert rs.classify_deterministic(
        "microsoftexchange329e71ec88ae@dominio.on", "Non recapitabile: Confronto", ""
    ) == rs.BOUNCE
    assert rs.classify_deterministic(
        "a@b.it", "qualsiasi",
        "Non è stato possibile recapitare il messaggio a x@y.it."
    ) == rs.BOUNCE


# --------------------------------------------------------------------------
# Classificatore euristico
# --------------------------------------------------------------------------
def test_euristica_rifiuto_ho_gia_consulente():
    out = rs.HeuristicClassifier().classify("La ringrazio ma ho già un consulente di fiducia.")
    assert out["label"] == rs.RIFIUTO           # NON altro (spec)
    assert out["confidence"] >= 0.7


def test_euristica_referente_con_email():
    out = rs.HeuristicClassifier().classify(
        "Può rivolgersi a Marco Longhi (marco.longhi@ente.it), se ne occupa lui."
    )
    assert out["label"] == rs.REFERENTE
    assert out["referente"]["email"] == "marco.longhi@ente.it"


def test_euristica_rifiuto_vince_su_positiva():
    # frase mista: il rifiuto deve prevalere (mai falso positivo su positiva)
    out = rs.HeuristicClassifier().classify(
        "Mi interessa il tema, ma ho già un consulente e non desidero altri contatti."
    )
    assert out["label"] == rs.RIFIUTO


def test_euristica_ambigua_bassa_confidenza():
    out = rs.HeuristicClassifier().classify("Inoltro ricevuto per conoscenza.")
    assert out["label"] == rs.ALTRO
    assert out["confidence"] < rs.CONFIDENCE_THRESHOLD


def test_euristica_referente_copia_collega():
    """Pattern dai casi reali: 'copio il collega … per indirizzarla al meglio'."""
    out = rs.HeuristicClassifier().classify(
        "Grazie per la mail, copio il collega Aldo Riva di HR al fine di "
        "indirizzarla al meglio."
    )
    assert out["label"] == rs.REFERENTE


def test_euristica_accettazione_riunione_da_oggetto():
    """Il segnale può stare solo nell'oggetto (corpo = solo banner EXTERNAL)."""
    out = rs.HeuristicClassifier().classify(
        "Oggetto: [EXTERNAL] Accettata: Consulenza previdenziale\n"
        "--- This message is from an EXTERNAL SENDER - be CAUTIOUS ---"
    )
    assert out["label"] == rs.POSITIVA


def test_euristica_disposizione_colloquio():
    out = rs.HeuristicClassifier().classify(
        "La ringrazio per la comunicazione. Resto sempre a disposizione per "
        "un colloquio."
    )
    assert out["label"] == rs.POSITIVA


# --------------------------------------------------------------------------
# Scan + linking
# --------------------------------------------------------------------------
def _registro_rows():
    return [
        {"_row": 2, "Email": "mario.rossi@ente.it", "Stato": "Nessuna risposta"},
        {"_row": 3, "Email": "luca.bianchi@ente.it", "Stato": "Nessuna risposta"},
    ]


def test_scan_match_diretto_e_bounce_dal_corpo():
    export = [
        {"tipo": "ricevuto", "mittente": "mario.rossi@ente.it",
         "oggetto": "RE: proposta", "corpo_email": "Mi interessa, sentiamoci.",
         "conversationId": "c1"},
        {"tipo": "ricevuto", "mittente": "postmaster@ente.it",
         "oggetto": "Undeliverable", "corpo_email":
         "Delivery failed: luca.bianchi@ente.it not found", "conversationId": "c2"},
        {"tipo": "inviato", "mittente": "io@me.it", "oggetto": "x",
         "corpo_email": "x", "conversationId": "c3"},                 # ignorato
        {"tipo": "ricevuto", "mittente": "sconosciuto@altro.it",
         "oggetto": "spam", "corpo_email": "ciao", "conversationId": "c4"},  # non legato
    ]
    results = rs.scan(export, _registro_rows(), rs.HeuristicClassifier())
    assert len(results) == 2
    by_email = {r["email"]: r for r in results}
    assert by_email["mario.rossi@ente.it"]["label"] == rs.POSITIVA
    assert by_email["luca.bianchi@ente.it"]["label"] == rs.BOUNCE
    assert by_email["luca.bianchi@ente.it"]["stato"] == "Mancato recapito"


# --------------------------------------------------------------------------
# apply_results: scritture E/G/H/F, mai I/J, idempotenza, dry-run
# --------------------------------------------------------------------------
def _make_master(tmp_path: Path) -> Path:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Registro invii"
    for c, h in enumerate(["Data invio", "Email", "Ente/dominio", "Oggetto", "Stato",
                           "Azione consigliata", "Risposta associata", "Preview risposta",
                           "Gg lav da invio", "Variante"], start=1):
        ws.cell(1, c, h)
    ws.cell(2, 2, "mario.rossi@ente.it")
    ws.cell(2, 5, "Nessuna risposta")
    ws.cell(2, 9, "=NETWORKDAYS(A2,TODAY())")
    ws.cell(2, 10, "A")
    path = tmp_path / "master.xlsx"
    wb.save(path)
    return path


def _one_result(writable=True):
    return [{
        "registro_row": {"_row": 2, "Email": "mario.rossi@ente.it"},
        "email": "mario.rossi@ente.it",
        "label": rs.POSITIVA,
        "stato": "Risposta positiva / referente",
        "confidence": 0.9,
        "preview": "Mi interessa, sentiamoci.",
        "referente": None,
        "sintesi": "Interessato a un confronto.",
        "writable": writable,
    }]


def test_apply_scrive_e_g_h_f_ma_non_i_j(tmp_path):
    master = _make_master(tmp_path)
    rs.apply_results(master, _one_result(), apply=True,
                     backup_dir=tmp_path / "b", logs_dir=tmp_path / "l")
    wb = openpyxl.load_workbook(master)
    ws = wb["Registro invii"]
    assert ws.cell(2, 5).value == "Risposta positiva / referente"   # E
    assert ws.cell(2, 7).value == "Sì"                              # G
    assert ws.cell(2, 8).value == "Mi interessa, sentiamoci."       # H
    assert "[scanner" in ws.cell(2, 6).value                        # F append
    assert str(ws.cell(2, 9).value).startswith("=")                 # I intatta
    assert ws.cell(2, 10).value == "A"                              # J intatta


def test_apply_idempotente(tmp_path):
    master = _make_master(tmp_path)
    rs.apply_results(master, _one_result(), apply=True,
                     backup_dir=tmp_path / "b", logs_dir=tmp_path / "l")
    first = openpyxl.load_workbook(master)["Registro invii"].cell(2, 6).value
    # secondo run identico → nessun doppio append né doppia scrittura
    log = rs.apply_results(master, _one_result(), apply=True,
                           backup_dir=tmp_path / "b", logs_dir=tmp_path / "l")
    second = openpyxl.load_workbook(master)["Registro invii"].cell(2, 6).value
    assert first == second
    assert len(log.writes) == 0


def test_dry_run_non_tocca_il_master(tmp_path):
    master = _make_master(tmp_path)
    before = master.read_bytes()
    rs.apply_results(master, _one_result(), apply=False,
                     backup_dir=tmp_path / "b", logs_dir=tmp_path / "l")
    assert master.read_bytes() == before


def test_bassa_confidenza_non_scrive(tmp_path):
    master = _make_master(tmp_path)
    res = _one_result(writable=False)
    log = rs.apply_results(master, res, apply=True,
                           backup_dir=tmp_path / "b", logs_dir=tmp_path / "l")
    wb = openpyxl.load_workbook(master)
    assert wb["Registro invii"].cell(2, 5).value == "Nessuna risposta"
    assert len(log.skipped) == 1                 # finito in DA RIVEDERE


# --------------------------------------------------------------------------
# Eval end-to-end sul set d'esempio → deve passare le soglie
# --------------------------------------------------------------------------
def test_eval_example_passa_le_soglie():
    from evals.eval_scanner import evaluate, load_jsonl

    path = Path(__file__).resolve().parent.parent / "evals" / "replies_labeled.example.jsonl"
    res = evaluate(load_jsonl(path), rs.HeuristicClassifier())
    assert res["accuracy"] >= 0.90, res["errors"]
    assert res["fp_positiva"] == [], res["fp_positiva"]
    assert res["pass"] is True
