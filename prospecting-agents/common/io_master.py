"""Lettura/scrittura SICURA del master Excel.

Regole non negoziabili implementate qui (spec § Regole trasversali #3, #6):
- Prima di ogni scrittura reale → backup `backup/master_YYYYMMDD_HHMMSS.xlsx`.
- Mai cancellare righe.
- Mai scrivere in celle con formule (colonne calcolate read-only).
- Dry-run di default: `save()` scrive solo con apply=True.
- Idempotenza: `set_cell()` non riscrive un valore già presente (ritorna False)
  e non tocca mai una colonna formula.
- Ogni run → log JSON in `logs/`.

Dipendenza unica: openpyxl (gratuita).
"""

from __future__ import annotations

import json
import shutil
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

# ---------------------------------------------------------------------------
# Colonne READ-ONLY (formule) per foglio — MAI scrivibili.
# Chiave = nome foglio (o prefisso, vedi _readonly_cols); valore = lettere col.
# ---------------------------------------------------------------------------
FORMULA_COLUMNS: dict[str, set[str]] = {
    "Registro invii": {"I"},                    # Gg lav da invio
    "Follow-up e riprese": {"I", "J", "K"},     # Gg lavorativi / Tipo azione / Template
    "Test A-B": {"C", "D", "E"},                # verdetto a 3 vie via formule
}

# Prefissi foglio con colonne formula (le code hanno nomi variabili).
FORMULA_COLUMNS_BY_PREFIX: dict[str, set[str]] = {
    # nelle code, Gg-lavorativi non esiste; nessuna formula da proteggere di default
}


class FormulaCellError(Exception):
    """Tentativo di scrivere in una colonna calcolata (read-only)."""


@dataclass
class RunLog:
    """Log JSON di un run, salvato in logs/."""

    component: str
    apply: bool
    started_at: str = field(default_factory=lambda: datetime.now().isoformat())
    backup_path: str | None = None
    writes: list[dict[str, Any]] = field(default_factory=list)
    skipped: list[dict[str, Any]] = field(default_factory=list)
    notes: list[str] = field(default_factory=list)

    def record_write(self, sheet: str, cell: str, old: Any, new: Any) -> None:
        self.writes.append({"sheet": sheet, "cell": cell, "old": _s(old), "new": _s(new)})

    def record_skip(self, sheet: str, cell: str, reason: str) -> None:
        self.skipped.append({"sheet": sheet, "cell": cell, "reason": reason})

    def to_dict(self) -> dict[str, Any]:
        return {
            "component": self.component,
            "apply": self.apply,
            "started_at": self.started_at,
            "finished_at": datetime.now().isoformat(),
            "backup_path": self.backup_path,
            "n_writes": len(self.writes),
            "n_skipped": len(self.skipped),
            "writes": self.writes,
            "skipped": self.skipped,
            "notes": self.notes,
        }

    def save(self, logs_dir: str | Path) -> Path:
        logs_dir = Path(logs_dir)
        logs_dir.mkdir(parents=True, exist_ok=True)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = logs_dir / f"{self.component}_{ts}.json"
        path.write_text(json.dumps(self.to_dict(), ensure_ascii=False, indent=2))
        return path


def _s(v: Any) -> Any:
    """Rende un valore serializzabile in JSON."""
    if isinstance(v, datetime):
        return v.isoformat()
    return v


# ---------------------------------------------------------------------------
# Colonne read-only
# ---------------------------------------------------------------------------
def readonly_cols(sheet_name: str) -> set[str]:
    """Colonne formula (read-only) per un dato foglio, per nome o prefisso."""
    cols = set(FORMULA_COLUMNS.get(sheet_name, set()))
    for prefix, pcols in FORMULA_COLUMNS_BY_PREFIX.items():
        if sheet_name.startswith(prefix):
            cols |= pcols
    return cols


def is_formula_column(sheet_name: str, col_letter: str) -> bool:
    return col_letter.upper() in readonly_cols(sheet_name)


# ---------------------------------------------------------------------------
# Backup
# ---------------------------------------------------------------------------
def backup_master(master_path: str | Path, backup_dir: str | Path = "backup") -> Path:
    """Copia il master in backup/master_YYYYMMDD_HHMMSS.xlsx. Ritorna il path."""
    master_path = Path(master_path)
    if not master_path.exists():
        raise FileNotFoundError(f"Master non trovato: {master_path}")
    backup_dir = Path(backup_dir)
    backup_dir.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    dest = backup_dir / f"master_{ts}.xlsx"
    shutil.copy2(master_path, dest)
    return dest


# ---------------------------------------------------------------------------
# Lettura
# ---------------------------------------------------------------------------
def load(master_path: str | Path, data_only: bool = False) -> openpyxl.Workbook:
    return openpyxl.load_workbook(master_path, data_only=data_only)


def read_rows(
    wb: openpyxl.Workbook, sheet: str, header_row: int = 1
) -> list[dict[str, Any]]:
    """Legge un foglio come lista di dict header->valore.

    Ogni dict include `_row` (numero riga Excel, 1-based) per scritture mirate.
    """
    ws = wb[sheet]
    headers: dict[int, str] = {}
    for c in range(1, ws.max_column + 1):
        h = ws.cell(header_row, c).value
        if h is not None:
            headers[c] = str(h).strip()
    rows: list[dict[str, Any]] = []
    for r in range(header_row + 1, ws.max_row + 1):
        row = {"_row": r}
        empty = True
        for c, name in headers.items():
            v = ws.cell(r, c).value
            row[name] = v
            if v is not None and str(v).strip() != "":
                empty = False
        if not empty:
            rows.append(row)
    return rows


# ---------------------------------------------------------------------------
# Scrittura sicura + idempotente
# ---------------------------------------------------------------------------
def set_cell(
    ws: Worksheet,
    row: int,
    col: str | int,
    value: Any,
    *,
    sheet_name: str | None = None,
    log: RunLog | None = None,
) -> bool:
    """Scrive una cella in modo sicuro e idempotente.

    - Rifiuta le colonne formula (FormulaCellError).
    - Se il valore corrente è già == value → NON scrive, ritorna False (idempotenza).
    - Altrimenti scrive e ritorna True.

    NB: modifica solo l'oggetto worksheet in memoria; la persistenza avviene con
    `save(..., apply=True)`.
    """
    col_letter = col if isinstance(col, str) else get_column_letter(col)
    col_letter = col_letter.upper()
    name = sheet_name or ws.title

    if is_formula_column(name, col_letter):
        raise FormulaCellError(
            f"Colonna calcolata read-only: {name}!{col_letter} "
            f"(vedi FORMULA_COLUMNS)"
        )

    cell = ws.cell(row, column_index_from_string(col_letter))
    # non scrivere sopra una formula reale, comunque
    if isinstance(cell.value, str) and cell.value.startswith("="):
        raise FormulaCellError(
            f"La cella {name}!{col_letter}{row} contiene una formula: {cell.value!r}"
        )

    if _equal(cell.value, value):
        if log:
            log.record_skip(name, f"{col_letter}{row}", "già a questo valore (idempotente)")
        return False

    old = cell.value
    cell.value = value
    if log:
        log.record_write(name, f"{col_letter}{row}", old, value)
    return True


def _equal(a: Any, b: Any) -> bool:
    if a is None and (b is None or b == ""):
        return True
    if isinstance(a, str) and isinstance(b, str):
        return a.strip() == b.strip()
    return a == b


def save(
    wb: openpyxl.Workbook,
    master_path: str | Path,
    *,
    apply: bool = False,
    backup_dir: str | Path = "backup",
    log: RunLog | None = None,
) -> Path | None:
    """Persiste il workbook SOLO se apply=True (dry-run di default).

    In modalità apply: esegue prima il backup del master ESISTENTE, poi salva.
    Ritorna il path del backup (o None in dry-run).
    """
    if not apply:
        if log:
            log.notes.append("DRY-RUN: nessuna scrittura su disco (usa --apply)")
        return None
    bkp = backup_master(master_path, backup_dir)
    if log:
        log.backup_path = str(bkp)
    wb.save(master_path)
    return bkp
